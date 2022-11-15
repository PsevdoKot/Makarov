import csv
import re
import os
from functools import cmp_to_key
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
import matplotlib.pyplot as plt
import numpy as np


class Vacancy:
    def __init__(self, name, description, key_skills, experience_id, premium,
                 employer_name, salary, area_name, published_at):
        self.name = name
        self.description = description
        self.key_skills = key_skills
        self.experience_id = experience_id
        self.premium = premium
        self.employer_name = employer_name
        self.salary = salary
        self.area_name = area_name
        self.published_at = published_at


class Salary:
    def __init__(self, salary_from, salary_to, salary_gross, salary_currency):
        self.salary_from = salary_from
        self.salary_to = salary_to
        self.salary_gross = salary_gross
        self.salary_currency = salary_currency

    def currency_to_rur(self):
        dic_currency_to_rub = {"Манаты": 35.68, "Белорусские рубли": 23.91, "Евро": 59.90, "Грузинский лари": 21.74,
                               "Киргизский сом": 0.76, "Тенге": 0.13, "Рубли": 1, "Гривны": 1.64, "Доллары": 60.66,
                               "Узбекский сум": 0.0055}
        return list(map(lambda x: int(x.replace(' ', '')) * dic_currency_to_rub[self.salary_currency],
                        (self.salary_from, self.salary_to)))

    def get_salary(self):
        return sum(self.currency_to_rur()) / 2


class DataSet:
    def __init__(self, file_name):
        (headers, info) = self._csv_reader(file_name)
        vacancies = self._csv_filter(headers, info) if len(headers) > 6 else self._small_csv_filter(info)
        self.file_name = file_name
        self.vacancies_objects = vacancies

    def _csv_reader(self, file_name):
        with open(file_name, encoding="utf-8-sig") as f:
            reader = [x for x in csv.reader(f)]
            headers = reader.pop(0)
            header_len = len(headers)
            info = list(filter(lambda data: '' not in data and len(data) == header_len, reader))
        return headers, info

    def _csv_filter(self, headers, info):
        def normalize_info_from_csv(info_cell):
            temp_info = "__temp__".join(info_cell.split("\n"))
            temp_info = re.sub(r"<[^<>]*>", "", temp_info)
            temp_info = re.sub(r"\s+", " ", temp_info)
            return str.strip(temp_info)

        vacancies = []
        for info_row in info:
            info_list = list(map(lambda x: normalize_info_from_csv(info_row[x]), range(len(headers))))
            salary = Salary(info_list[6], info_list[7], info_list[8], info_list[9])
            key_skills = info_list[2].split('__temp__')
            vacancy = Vacancy(info_list[0], info_list[1], key_skills, info_list[3], info_list[4],
                              info_list[5], salary, info_list[10], info_list[11])
            vacancies.append(vacancy)
        return vacancies

    def _small_csv_filter(self, info):
        vacancies = []
        for info_row in info:
            salary = Salary(info_row[1], info_row[2], None, info_row[3])
            vacancies.append(Vacancy(info_row[0], None, None, None, None, None, salary, info_row[4], info_row[5]))
        return vacancies


class InputConnect:
    def info_formatter(self, vacancies):
        def formatter_string_number(str_num):
            return str_num if str_num.find('.') == -1 else str_num[:len(str_num) - 2]

        def formatter_salary(attr_value):
            salary_from = formatter_string_number(attr_value.salary_from)
            salary_to = formatter_string_number(attr_value.salary_to)
            salary_currency = dic_currency[attr_value.salary_currency]
            return Salary(salary_from, salary_to, None, salary_currency)

        def formatter_published_at(attr_value):
            return attr_value[0:4]

        dic_currency = {"AZN": "Манаты", "BYR": "Белорусские рубли", "EUR": "Евро",
                        "GEL": "Грузинский лари", "KGS": "Киргизский сом", "KZT": "Тенге", "RUR": "Рубли",
                        "UAH": "Гривны", "USD": "Доллары", "UZS": "Узбекский сум"}

        for vacancy in vacancies:
            setattr(vacancy, "salary", formatter_salary(getattr(vacancy, "salary")))
            setattr(vacancy, "published_at", formatter_published_at(getattr(vacancy, "published_at")))
        return vacancies

    def info_finder(self, vacancies, finder_parameter):
        salaries_year_level, selected_salary_year_level, vacancies_year_count, selected_vacancy_year_count, \
        salaries_city_level, vacancies_city_count = {}, {}, {}, {}, {}, {}
        for vacancy in vacancies:
            salary = vacancy.salary.get_salary()
            year = int(vacancy.published_at)
            if year not in salaries_year_level:
                salaries_year_level[year] = (salary, 1)
                vacancies_year_count[year] = 1
                selected_salary_year_level[year] = (0, 0)
                selected_vacancy_year_count[year] = 0
            else:
                sal_yr_lvl = salaries_year_level[year]
                salaries_year_level[year] = (sal_yr_lvl[0] + salary, sal_yr_lvl[1] + 1)
                vacancies_year_count[year] += 1
            if finder_parameter in vacancy.name:
                sel_sal_ye_lvl = selected_salary_year_level[year]
                selected_salary_year_level[year] = (sel_sal_ye_lvl[0] + salary, sel_sal_ye_lvl[1] + 1)
                selected_vacancy_year_count[year] += 1
            if vacancy.area_name not in salaries_city_level:
                vacancies_city_count[vacancy.area_name] = 1
                salaries_city_level[vacancy.area_name] = (salary, 1)
            else:
                sal_ct_lvl = salaries_city_level[vacancy.area_name]
                salaries_city_level[vacancy.area_name] = (sal_ct_lvl[0] + salary, sal_ct_lvl[1] + 1)
                vacancies_city_count[vacancy.area_name] += 1
        return self._info_calculating(salaries_year_level, selected_salary_year_level, vacancies_year_count,
                                      selected_vacancy_year_count, salaries_city_level, vacancies_city_count, len(vacancies))

    def _info_calculating(self, salaries_year_level, selected_salary_year_level, vacancies_year_count,
                          selected_vacancy_year_count, salaries_city_level, vacancies_city_count, vacancies_count):
        def sort_dict(dictionary):
            dict_pairs = [(key, value) for key, value in dictionary.items()]
            dict_pairs.sort(key=cmp_to_key(lambda x, y: -1 if x[1] <= y[1] else 1))
            return dict(dict_pairs)

        (salaries_year_level, selected_salary_year_level, salaries_city_level) = \
            list(map(lambda dictionary:
                dict(map(lambda dict_pair:
                    (dict_pair[0], int(dict_pair[1][0] / dict_pair[1][1]) if dict_pair[1][1] != 0 else int(dict_pair[1][0])), dictionary.items())),
                (salaries_year_level, selected_salary_year_level, salaries_city_level)))
        vacancies_city_count = dict(map(lambda dict_pair: (dict_pair[0], float(f"{dict_pair[1] / vacancies_count:.4f}")), vacancies_city_count.items()))
        vacancies_city_count = dict(filter(lambda dict_pair: dict_pair[1] >= 0.01, vacancies_city_count.items()))
        vacancies_city_count = sort_dict(vacancies_city_count)
        vacancies_city_count = {k: vacancies_city_count[k] for k in list(vacancies_city_count)[-10:][::-1]}
        vacancies_city_count = dict(map(lambda dict_pair: (dict_pair[0], f"{round(dict_pair[1] * 100, 2)}%"), vacancies_city_count.items()))
        salaries_city_level = dict(filter(lambda dict_pair: dict_pair[0] in vacancies_city_count, salaries_city_level.items()))
        salaries_city_level = sort_dict(salaries_city_level)
        salaries_city_level = {k: salaries_city_level[k] for k in list(salaries_city_level)[-10:][::-1]}
        return salaries_year_level, selected_salary_year_level, vacancies_year_count, \
               selected_vacancy_year_count, salaries_city_level, vacancies_city_count, vacancies_count


class Report:
    def __init__(self, vacancy_info):
        self.salaries_year_level = vacancy_info[0]
        self.vacancies_year_count = vacancy_info[1]
        self.selected_salary_year_level = vacancy_info[2]
        self.selected_vacancy_year_count = vacancy_info[3]
        self.salaries_city_level = vacancy_info[4]
        self.vacancies_city_count = vacancy_info[5]

    def generate_excel(self, vacancy_name):
        wb = Workbook()
        stats_by_year = wb.worksheets[0]
        stats_by_year.title = "Cтатистика по годам"
        stats_by_city = wb.create_sheet("Cтатистика по городам")

        stats_by_year.append(["Год", "Средняя зарплата", f"Средняя зарплата - {vacancy_name}",
                              "Количество вакансий", f"Количество вакансий - {vacancy_name}"])
        for i, year in enumerate(self.salaries_year_level.keys(), 2):
            stats_by_year.cell(row=i, column=1, value=year)
            for j, dictionary in enumerate((self.salaries_year_level, self.vacancies_year_count,
                                            self.selected_salary_year_level, self.selected_vacancy_year_count), 2):
                stats_by_year.cell(row=i, column=j, value=dictionary[year])

        stats_by_city.append(["Город", "Уровень зарплат", "", "Город", "Доля вакансий"])
        for i, city in enumerate(self.salaries_city_level.keys(), 2):
            stats_by_city.cell(row=i, column=1, value=city)
            stats_by_city.cell(row=i, column=2, value=self.salaries_city_level[city])
        for i, city in enumerate(self.vacancies_city_count.keys(), 2):
            stats_by_city.cell(row=i, column=4, value=city)
            stats_by_city.cell(row=i, column=5, value=self.vacancies_city_count[city])

        self._slylize_wb(wb)
        wb.save('report.xlsx')

    def _slylize_wb(self, wb):
        bold_font = Font(bold=True)
        thin = Side(border_style="thin", color="000000")
        outline = Border(top=thin, left=thin, right=thin, bottom=thin)
        for worksheet in wb.worksheets:
            for column_cells in worksheet.columns:
                length = max(len(str(cell.value) if cell.value is not None else "") for cell in column_cells)
                worksheet.column_dimensions[column_cells[0].column_letter].width = length + 3
            for cell in worksheet[1]:
                cell.font = bold_font
            for column in tuple(worksheet.columns):
                if column[1].value == None:
                    continue
                for cell in column:
                    cell.border = outline

    def generate_image(self, vacancy_name):
        fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(2, 2, figsize=(12, 7.5), layout='constrained')
        self._generate_salary_year_levels_graph(ax1, vacancy_name)
        self._generate_vacancy_year_count_graph(ax2, vacancy_name)
        self._generate_salary_city_levels_graph(ax3)
        self._generate_vacancy_city_count_graph(ax4)
        plt.savefig('graph.png')

    def _generate_salary_year_levels_graph(self, ax, vacancy_name):
        ax_labels = self.salaries_year_level.keys()
        x = np.arange(len(ax_labels))
        width = 0.35
        ax.bar(x - width / 2, self.salaries_year_level.values(), width, label='Средняя з/п')
        ax.bar(x + width / 2, self.selected_salary_year_level.values(), width, label=f'З/п {vacancy_name}')
        ax.set_xticks(x, ax_labels, fontsize=8, rotation=90, ha='right')
        ax.set_title("Уровень зарплат по годам")
        ax.yaxis.grid(True)
        ax.legend(fontsize=8, loc='upper left')

    def _generate_vacancy_year_count_graph(self, ax, vacancy_name):
        ax_labels = self.vacancies_year_count.keys()
        x = np.arange(len(ax_labels))
        width = 0.35
        ax.bar(x - width / 2, self.vacancies_year_count.values(), width, label='Количество вакансий')
        ax.bar(x + width / 2, self.selected_vacancy_year_count.values(), label=f'Количество вакансий {vacancy_name}')
        ax.set_xticks(x, ax_labels, fontsize=8, rotation=90, ha='right')
        ax.set_title("Количество вакансий по годам")
        ax.yaxis.grid(True)
        ax.legend(fontsize=8, loc='upper left')

    def _generate_salary_city_levels_graph(self, ax):
        ax_labels = list(
            map(lambda value: value.replace('-', '-\n'), self.salaries_city_level.keys()))
        y_pos = np.arange(len(ax_labels))
        ax.barh(y_pos, self.salaries_city_level.values(), align='center')
        ax.set_yticks(y_pos, fontsize=6, labels=ax_labels)
        ax.invert_yaxis()
        ax.set_title("Уровень зарплат по городам")

    def _generate_vacancy_city_count_graph(self, ax):
        ax_labels, values = list(self.vacancies_city_count.keys()), self.vacancies_city_count.values()
        ax_labels.append('Другие')
        values = list(map(lambda value: float(value[:-1]), values))
        values.append(100 - sum(values))
        ax.pie(values, labels=ax_labels)
        ax.set_title("Доля вакансий по городам")


def normalize_input_info(input_info):
    if os.stat(input_info[0]).st_size == 0:
        return "Пустой файл"
    return "Нормализация прошла успешно"


def main_function():
    input_requests = ["Введите название файла: ", "Введите название профессии: "]
    input_info = [input(input_request) for input_request in input_requests]  # vacancies из условия == vacancies_by_year
    # input_info = ["vacancies_by_year.csv", "аналитик"]
    normalize_result = normalize_input_info(input_info)
    if normalize_result != "Нормализация прошла успешно":
        return normalize_result
    data_set = DataSet(input_info[0])
    if len(data_set.vacancies_objects) == 0:
        return "Нет данных"
    input_connect = InputConnect()
    formatted_info = input_connect.info_formatter(data_set.vacancies_objects)
    info = input_connect.info_finder(formatted_info, input_info[1])
    report = Report(info)
    report.generate_excel(input_info[1])
    report.generate_image(input_info[1])


main_function()

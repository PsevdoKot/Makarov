import csv
import re
import os
from functools import cmp_to_key
from prettytable import PrettyTable
from prettytable import ALL
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
import matplotlib.pyplot as plt
import numpy as np
from jinja2 import Environment, FileSystemLoader
import pdfkit


def normalize_input_info(input_info):
    table_fields = ["Название", "Описание", "Навыки", "Опыт работы", "Премиум-вакансия", "Компания", "Оклад",
                    "Название региона", "Дата публикации вакансии", "Идентификатор валюты оклада", "None"]
    if os.stat(input_info[0]).st_size == 0:
        return "Пустой файл"
    if input_info[1] == '':
        input_info[1] = "None: None"
    temp = input_info[1].find(': ')
    if temp == -1:
        return "Формат ввода некорректен"
    input_info[1] = [input_info[1][:temp], input_info[1][temp + 2:]]
    if input_info[1][0] not in table_fields:
        return "Параметр поиска некорректен"
    if input_info[2] == '':
        input_info[2] = '№'
    elif input_info[2] not in table_fields:
        return "Параметр сортировки некорректен"
    if input_info[3] == 'Да':
        input_info[3] = True
    elif input_info[3] == 'Нет' or input_info[3] == '':
        input_info[3] = False
    else:
        return "Порядок сортировки задан некорректно"
    input_info[4] = list(map(lambda x: 0 if x == '' else int(x) - 1, input_info[4].split(' ')))
    if len(input_info[4]) == 1:
        input_info[4].append(10000)
    input_info[5] = input_info[5].split(', ')
    if input_info[5][0] == '':
        input_info[5] = table_fields[:9]
    input_info[5].insert(0, '№')
    return "Нормализация прошла успешно"


def csv_reader(file_name):
    with open(file_name, encoding="utf-8-sig") as f:
        reader = [x for x in csv.reader(f)]
        headers = reader.pop(0)
        header_len = len(headers)
        info = list(filter(lambda data: '' not in data and len(data) == header_len, reader))
    return (headers, info)


def csv_filter(headers, info):
    def normalize_info_from_csv(info_cell):
        temp_info = "__temp__".join(info_cell.split("\n"))
        temp_info = re.sub(r"<[^<>]*>", "", temp_info)
        temp_info = re.sub(r"\s+", " ", temp_info)
        return str.strip(temp_info)

    info_dictionaries = []
    for info_row in info:
        info_dictionary = {}
        for i in range(len(headers)):
            info_dictionary[headers[i]] = normalize_info_from_csv(info_row[i])
        info_dictionaries.append(info_dictionary)
    return info_dictionaries


def info_formatter(info_dictionaries):
    def formatter_string_number(str_num):
        num = int(str_num if str_num.find('.') == -1 else str_num[:len(str_num) - 2])
        str_num_reverse = str(num)[::-1]
        return ' '.join(str_num_reverse[i:i + 3] for i in range(0, len(str_num_reverse), 3))[::-1]

    def formatter_experience_id(new_info_dictionary, value, key):
        new_info_dictionary["Опыт работы"] = dic_experience[value]

    def formatter_salary_from(new_info_dictionary, value, key):
        new_info_dictionary['Оклад'] = formatter_string_number(value)

    def formatter_salary_to(new_info_dictionary, value, key):
        new_info_dictionary['Оклад'] = f"{new_info_dictionary['Оклад']} - {formatter_string_number(value)}"

    def formatter_salary_currency(new_info_dictionary, value, key):
        new_info_dictionary["Оклад"] = f"{new_info_dictionary['Оклад']} ({dic_currency[value]}) ({new_info_dictionary['salary_currency']})"

    def formatter_salary_gross(new_info_dictionary, value, key):
        new_info_dictionary['salary_currency'] = 'Без вычета налогов' if value == 'True' else 'С вычетом налогов' if value == 'False' else value

    def formatter_published_at(new_info_dictionary, value, key):
        new_info_dictionary["Дата публикации вакансии"] = f"{value}#{value[8:10]}.{value[5:7]}.{value[0:4]}"

    def formatter_premium(new_info_dictionary, value, key):
        new_info_dictionary["Премиум-вакансия"] = 'Да' if value == 'True' else 'Нет'

    def formatter_key_skills(new_info_dictionary, value, key):
        value = value.replace("__temp__", '\n')
        new_info_dictionary["Количество навыков"] = value.count('\n') + 1
        new_info_dictionary["Навыки"] = f"{value[0:100]}..." if len(value) > 100 else value

    def formatter_standart_field_value(new_info_dictionary, value, key):
        new_info_dictionary[dic_naming[key]] = f"{value[0:100]}..." if len(value) > 100 else value

    dic_naming = {"name": "Название", "description": "Описание", "employer_name": "Компания",
                  "area_name": "Название региона"}
    dic_experience = {"noExperience": "Нет опыта", "between1And3": "От 1 года до 3 лет",
                      "between3And6": "От 3 до 6 лет", "moreThan6": "Более 6 лет"}
    dic_currency = {"AZN": "Манаты", "BYR": "Белорусские рубли", "EUR": "Евро",
                    "GEL": "Грузинский лари", "KGS": "Киргизский сом", "KZT": "Тенге", "RUR": "Рубли",
                    "UAH": "Гривны", "USD": "Доллары", "UZS": "Узбекский сум"}
    dic_func = {"experience_id": formatter_experience_id, "salary_from": formatter_salary_from,
                "salary_to": formatter_salary_to, "salary_currency": formatter_salary_currency,
                "salary_gross": formatter_salary_gross, "published_at": formatter_published_at,
                "premium": formatter_premium, "key_skills": formatter_key_skills,
                "name": formatter_standart_field_value, "description": formatter_standart_field_value,
                "employer_name": formatter_standart_field_value, "area_name": formatter_standart_field_value}

    formatted_info_dictionaries = []
    for info_dictionary in info_dictionaries:
        formatted_info_dictionary = {}
        for item_key, item_value in info_dictionary.items():
            dic_func[item_key](formatted_info_dictionary, item_value, item_key)
        formatted_info_dictionary.pop('salary_currency')
        formatted_info_dictionaries.append(formatted_info_dictionary)
    return formatted_info_dictionaries


def info_filter(info_dictionaries, filtering_parameter):
    def filter_verbatim(dic, field_value_should):
        return dic[field_value_should[0]] == field_value_should[1]

    def filter_key_skills(dic, filtering_parameter):
        field_value_should = filtering_parameter[1].split(', ')
        dic_values = dic['Навыки'].replace(', ', '\n').replace('...', '\n').split('\n')
        return all(list(map(lambda value_should: value_should in dic_values, field_value_should)))

    def filter_salary(dic, field_value_should):
        dic_value = dic["Оклад"]
        salary_area = dic_value[:dic_value.find('(')].replace(' ', '').split('-')
        return int(salary_area[0]) <= int(field_value_should[1]) <= int(salary_area[1])

    def filter_salary_currency(dic, field_value_should):
        dic_value = dic["Оклад"]
        temp = dic_value[dic_value.find('(') + 1:dic_value.find(')')]
        return temp == field_value_should[1]

    def filter_published_at(dic, field_value_should):
        dic_value = dic["Дата публикации вакансии"]
        return dic_value[dic_value.find('#') + 1:] == field_value_should[1]

    dic_filter = {"Название": filter_verbatim, "Описание": filter_verbatim, "Навыки": filter_key_skills,
                  "Опыт работы": filter_verbatim, "Премиум-вакансия": filter_verbatim, "Компания": filter_verbatim,
                  "Оклад": filter_salary, "Дата публикации вакансии": filter_published_at,
                  "Идентификатор валюты оклада": filter_salary_currency, "Название региона": filter_verbatim}

    return list(filter(lambda info_dictionary:
        filtering_parameter[0] == "None" or dic_filter[filtering_parameter[0]](info_dictionary, filtering_parameter),
        info_dictionaries))


def info_sorter(info_dictionaries, sort_field, reverse_sort):
    def lexcographic_sorter(row1, row2):
        return 1 if row1[sort_field] >= row2[sort_field] else -1

    def key_skills_sorter(row1, row2):
        (row1_len, row2_len) = list(map(lambda row: row["Количество навыков"], (row1, row2)))
        return row1_len - row2_len

    def experience_sorter(row1, row2):
        def find_first_num(row):
            row_value = row["Опыт работы"]
            row_num = list(filter(lambda char: char.isdigit(), row_value))
            return int(row_num[0]) if len(row_num) > 0 else 0
        (row1_num, row2_num) = list(map(lambda row: find_first_num(row), (row1, row2)))
        return row1_num - row2_num

    def salary_sorter(row1, row2):
        def salary_process(row):
            dic_currency_to_rub = {"Манаты": 35.68, "Белорусские рубли": 23.91, "Евро": 59.90, "Грузинский лари": 21.74,
                                   "Киргизский сом": 0.76, "Тенге": 0.13, "Рубли": 1, "Гривны": 1.64, "Доллары": 60.66,
                                   "Узбекский сум": 0.0055}
            string_nums = row[sort_field].split(' - ')
            salary_currency = row[sort_field][row[sort_field].find('(') + 1:row[sort_field].find(')')]
            string_nums[1] = string_nums[1][:string_nums[1].find(' (')]
            nums = list(map(lambda string_num: int(string_num.replace(' ', '')), string_nums))
            rub_nums = list(map(lambda num: num * dic_currency_to_rub[salary_currency], nums))
            return sum(rub_nums) / 2
        (row1_salary, row2_salary) = list(map(lambda row: salary_process(row), (row1, row2)))
        return row1_salary - row2_salary

    dic_sorter = {"Название": lexcographic_sorter, "Описание": lexcographic_sorter, "Навыки": key_skills_sorter,
                  "Опыт работы": experience_sorter, "Премиум-вакансия": lexcographic_sorter,
                  "Компания": lexcographic_sorter, "Оклад": salary_sorter, "Название региона": lexcographic_sorter,
                  "Дата публикации вакансии": lexcographic_sorter}

    info_dictionaries.sort(key=cmp_to_key(dic_sorter[sort_field]), reverse=reverse_sort)

    return info_dictionaries


def print_vacancies(info_dictionaries, start_end_nums, table_fields):
    info_table = PrettyTable(["Название", "Описание", "Навыки", "Опыт работы", "Премиум-вакансия",
                              "Компания", "Оклад", "Название региона", "Дата публикации вакансии"])
    for info_dictionary in info_dictionaries:
        values = list(map(lambda key: info_dictionary[key], info_dictionary))
        values.pop(2)
        info_table.add_row(values)
    published_at_data = list(filter(lambda x: x != '', info_table.get_string(
        fields=["Дата публикации вакансии"], border=False, header=False).replace(' ', '').split('\n')))
    info_table.del_column("Дата публикации вакансии")
    info_table.add_column("Дата публикации вакансии", list(map(lambda x: x[x.find('#') + 1:], published_at_data)))
    info_table.add_autoindex('№')
    info_table.hrules = ALL
    info_table.align = 'l'
    info_table.max_width = 20
    print(info_table.get_string(start=start_end_nums[0], end=start_end_nums[1], fields=table_fields))


#####################################################################################################################################


class Vacancy:
    def __init__(self, name, description, key_skills, experience_id, premium,
                 employer_name, salary, area_name, published_at):
        self.name = name
        self.description = description
        self.key_skills = key_skills
        self.experience_id = experience_id
        self.premium = premium
        self.employer_name = employer_name
        self.salary = salary
        self.area_name = area_name
        self.published_at = published_at


class Salary:
    def __init__(self, salary_from, salary_to, salary_gross, salary_currency):
        self.salary_from = salary_from
        self.salary_to = salary_to
        self.salary_gross = salary_gross
        self.salary_currency = salary_currency

    def currency_to_rur(self):
        dic_currency_to_rub = {"Манаты": 35.68, "Белорусские рубли": 23.91, "Евро": 59.90, "Грузинский лари": 21.74,
                               "Киргизский сом": 0.76, "Тенге": 0.13, "Рубли": 1, "Гривны": 1.64, "Доллары": 60.66,
                               "Узбекский сум": 0.0055}
        return list(map(lambda x: int(x.replace(' ', '')) * dic_currency_to_rub[self.salary_currency],
                        (self.salary_from, self.salary_to)))

    def get_salary(self):
        return sum(self.currency_to_rur()) / 2


class DataSet:
    def __init__(self, file_name):
        (headers, info) = self._csv_reader(file_name)
        vacancies = self._csv_filter(headers, info) if len(headers) > 6 else self._small_csv_filter(info)
        self.file_name = file_name
        self.vacancies_objects = vacancies

    def _csv_reader(self, file_name):
        with open(file_name, encoding="utf-8-sig") as f:
            reader = [x for x in csv.reader(f)]
            headers = reader.pop(0)
            header_len = len(headers)
            info = list(filter(lambda data: '' not in data and len(data) == header_len, reader))
        return headers, info

    def _csv_filter(self, headers, info):
        def normalize_info_from_csv(info_cell):
            temp_info = "__temp__".join(info_cell.split("\n"))
            temp_info = re.sub(r"<[^<>]*>", "", temp_info)
            temp_info = re.sub(r"\s+", " ", temp_info)
            return str.strip(temp_info)

        vacancies = []
        for info_row in info:
            info_list = list(map(lambda x: normalize_info_from_csv(info_row[x]), range(len(headers))))
            salary = Salary(info_list[6], info_list[7], info_list[8], info_list[9])
            key_skills = info_list[2].split('__temp__')
            vacancy = Vacancy(info_list[0], info_list[1], key_skills, info_list[3], info_list[4],
                              info_list[5], salary, info_list[10], info_list[11])
            vacancies.append(vacancy)
        return vacancies

    def _small_csv_filter(self, info):
        vacancies = []
        for info_row in info:
            salary = Salary(info_row[1], info_row[2], None, info_row[3])
            vacancies.append(Vacancy(info_row[0], None, None, None, None, None, salary, info_row[4], info_row[5]))
        return vacancies


class InputConnect:
    def info_formatter(self, vacancies):
        def formatter_string_number(str_num):
            return str_num if str_num.find('.') == -1 else str_num[:len(str_num) - 2]

        def formatter_salary(attr_value):
            salary_from = formatter_string_number(attr_value.salary_from)
            salary_to = formatter_string_number(attr_value.salary_to)
            salary_currency = dic_currency[attr_value.salary_currency]
            return Salary(salary_from, salary_to, None, salary_currency)

        def formatter_published_at(attr_value):
            return attr_value[0:4]

        dic_currency = {"AZN": "Манаты", "BYR": "Белорусские рубли", "EUR": "Евро",
                        "GEL": "Грузинский лари", "KGS": "Киргизский сом", "KZT": "Тенге", "RUR": "Рубли",
                        "UAH": "Гривны", "USD": "Доллары", "UZS": "Узбекский сум"}

        for vacancy in vacancies:
            setattr(vacancy, "salary", formatter_salary(getattr(vacancy, "salary")))
            setattr(vacancy, "published_at", formatter_published_at(getattr(vacancy, "published_at")))
        return vacancies

    def info_finder(self, vacancies, finder_parameter):
        salaries_year_level, selected_salary_year_level, vacancies_year_count, selected_vacancy_year_count, \
        salaries_city_level, vacancies_city_count = {}, {}, {}, {}, {}, {}
        for vacancy in vacancies:
            salary = vacancy.salary.get_salary()
            year = int(vacancy.published_at)
            if year not in salaries_year_level:
                salaries_year_level[year] = (salary, 1)
                vacancies_year_count[year] = 1
                selected_salary_year_level[year] = (0, 0)
                selected_vacancy_year_count[year] = 0
            else:
                sal_yr_lvl = salaries_year_level[year]
                salaries_year_level[year] = (sal_yr_lvl[0] + salary, sal_yr_lvl[1] + 1)
                vacancies_year_count[year] += 1
            if finder_parameter in vacancy.name:
                sel_sal_ye_lvl = selected_salary_year_level[year]
                selected_salary_year_level[year] = (sel_sal_ye_lvl[0] + salary, sel_sal_ye_lvl[1] + 1)
                selected_vacancy_year_count[year] += 1
            if vacancy.area_name not in salaries_city_level:
                vacancies_city_count[vacancy.area_name] = 1
                salaries_city_level[vacancy.area_name] = (salary, 1)
            else:
                sal_ct_lvl = salaries_city_level[vacancy.area_name]
                salaries_city_level[vacancy.area_name] = (sal_ct_lvl[0] + salary, sal_ct_lvl[1] + 1)
                vacancies_city_count[vacancy.area_name] += 1
        return self._info_calculating(salaries_year_level, selected_salary_year_level, vacancies_year_count,
                                      selected_vacancy_year_count, salaries_city_level, vacancies_city_count, len(vacancies))

    def _info_calculating(self, salaries_year_level, selected_salary_year_level, vacancies_year_count,
                          selected_vacancy_year_count, salaries_city_level, vacancies_city_count, vacancies_count):
        def sort_dict(dictionary):
            dict_pairs = [(key, value) for key, value in dictionary.items()]
            dict_pairs.sort(key=cmp_to_key(lambda x, y: -1 if x[1] <= y[1] else 1))
            return dict(dict_pairs)

        (salaries_year_level, selected_salary_year_level, salaries_city_level) = \
            list(map(lambda dictionary:
                dict(map(lambda dict_pair:
                    (dict_pair[0], int(dict_pair[1][0] / dict_pair[1][1]) if dict_pair[1][1] != 0 else int(dict_pair[1][0])), dictionary.items())),
                (salaries_year_level, selected_salary_year_level, salaries_city_level)))
        vacancies_city_count = dict(map(lambda dict_pair: (dict_pair[0], float(f"{dict_pair[1] / vacancies_count:.4f}")), vacancies_city_count.items()))
        vacancies_city_count = dict(filter(lambda dict_pair: dict_pair[1] >= 0.01, vacancies_city_count.items()))
        vacancies_city_count = sort_dict(vacancies_city_count)
        vacancies_city_count = {k: vacancies_city_count[k] for k in list(vacancies_city_count)[-10:][::-1]}
        vacancies_city_count = dict(map(lambda dict_pair: (dict_pair[0], f"{round(dict_pair[1] * 100, 2)}%"), vacancies_city_count.items()))
        salaries_city_level = dict(filter(lambda dict_pair: dict_pair[0] in vacancies_city_count, salaries_city_level.items()))
        salaries_city_level = sort_dict(salaries_city_level)
        salaries_city_level = {k: salaries_city_level[k] for k in list(salaries_city_level)[-10:][::-1]}
        return salaries_year_level, selected_salary_year_level, vacancies_year_count, \
               selected_vacancy_year_count, salaries_city_level, vacancies_city_count, vacancies_count


class Report:
    def __init__(self, vacancy_info):
        self.salaries_year_level = vacancy_info[0]
        self.vacancies_year_count = vacancy_info[1]
        self.selected_salary_year_level = vacancy_info[2]
        self.selected_vacancy_year_count = vacancy_info[3]
        self.salaries_city_level = vacancy_info[4]
        self.vacancies_city_count = vacancy_info[5]

    def generate_excel(self, vacancy_name):
        wb = Workbook()
        stats_by_year = wb.worksheets[0]
        stats_by_year.title = "Cтатистика по годам"
        stats_by_city = wb.create_sheet("Cтатистика по городам")

        stats_by_year.append(["Год", "Средняя зарплата", f"Средняя зарплата - {vacancy_name}",
                              "Количество вакансий", f"Количество вакансий - {vacancy_name}"])
        for i, year in enumerate(self.salaries_year_level.keys(), 2):
            stats_by_year.cell(row=i, column=1, value=year)
            for j, dictionary in enumerate((self.salaries_year_level, self.vacancies_year_count,
                                            self.selected_salary_year_level, self.selected_vacancy_year_count), 2):
                stats_by_year.cell(row=i, column=j, value=dictionary[year])

        stats_by_city.append(["Город", "Уровень зарплат", "", "Город", "Доля вакансий"])
        for i, city in enumerate(self.salaries_city_level.keys(), 2):
            stats_by_city.cell(row=i, column=1, value=city)
            stats_by_city.cell(row=i, column=2, value=self.salaries_city_level[city])
        for i, city in enumerate(self.vacancies_city_count.keys(), 2):
            stats_by_city.cell(row=i, column=4, value=city)
            stats_by_city.cell(row=i, column=5, value=self.vacancies_city_count[city])

        self._slylize_wb(wb)
        wb.save('report.xlsx')

    def _slylize_wb(self, wb):
        bold_font = Font(bold=True)
        thin = Side(border_style="thin", color="000000")
        outline = Border(top=thin, left=thin, right=thin, bottom=thin)
        for worksheet in wb.worksheets:
            for column_cells in worksheet.columns:
                length = max(len(str(cell.value) if cell.value is not None else "") for cell in column_cells)
                worksheet.column_dimensions[column_cells[0].column_letter].width = length + 3
            for cell in worksheet[1]:
                cell.font = bold_font
            for column in tuple(worksheet.columns):
                if column[1].value == None:
                    continue
                for cell in column:
                    cell.border = outline

    def generate_image(self, vacancy_name):
        fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(2, 2, figsize=(12, 7.5), layout='constrained')
        self._generate_salary_year_levels_graph(ax1, vacancy_name)
        self._generate_vacancy_year_count_graph(ax2, vacancy_name)
        self._generate_salary_city_levels_graph(ax3)
        self._generate_vacancy_city_count_graph(ax4)
        plt.savefig('graph.png')

    def _generate_salary_year_levels_graph(self, ax, vacancy_name):
        ax_labels = self.salaries_year_level.keys()
        x = np.arange(len(ax_labels))
        width = 0.35
        ax.bar(x - width / 2, self.salaries_year_level.values(), width, label='Средняя з/п')
        ax.bar(x + width / 2, self.selected_salary_year_level.values(), width, label=f'З/п {vacancy_name}')
        ax.set_xticks(x, ax_labels, fontsize=8, rotation=90, ha='right')
        ax.set_title("Уровень зарплат по годам")
        ax.yaxis.grid(True)
        ax.legend(fontsize=8, loc='upper left')

    def _generate_vacancy_year_count_graph(self, ax, vacancy_name):
        ax_labels = self.vacancies_year_count.keys()
        x = np.arange(len(ax_labels))
        width = 0.35
        ax.bar(x - width / 2, self.vacancies_year_count.values(), width, label='Количество вакансий')
        ax.bar(x + width / 2, self.selected_vacancy_year_count.values(), label=f'Количество вакансий {vacancy_name}')
        ax.set_xticks(x, ax_labels, fontsize=8, rotation=90, ha='right')
        ax.set_title("Количество вакансий по годам")
        ax.yaxis.grid(True)
        ax.legend(fontsize=8, loc='upper left')

    def _generate_salary_city_levels_graph(self, ax):
        ax_labels = self.salaries_city_level.keys()
        y_pos = np.arange(len(ax_labels))
        ax.barh(y_pos, self.salaries_city_level.values(), align='center')
        ax.set_yticks(y_pos, fontsize=8, labels=ax_labels)
        ax.invert_yaxis()
        ax.set_title("Уровень зарплат по городам")

    def _generate_vacancy_city_count_graph(self, ax):
        ax_labels, values = list(self.vacancies_city_count.keys()), self.vacancies_city_count.values()
        ax_labels.append('Другие')
        values = list(map(lambda value: float(value[:-1]), values))
        values.append(100 - sum(values))
        ax.pie(values, labels=ax_labels)
        ax.set_title("Доля вакансий по городам")

    def generate_pdf(self, vacancy_name):
        headers1, headers2, headers3 = (["Год", "Средняя зарплата", f"Средняя зарплата - {vacancy_name}",
                                        "Количество вакансий", f"Количество вакансий - {vacancy_name}"],
                                        ["Город", "Уровень зарплат"], ["Город", "Доля вакансий"])
        rows1 = list(map(lambda year: [year] + [dictionary[year] for dictionary in
                                       (self.salaries_year_level, self.vacancies_year_count,
                                        self.selected_salary_year_level, self.selected_vacancy_year_count)]
                         , self.salaries_year_level.keys()))
        rows2 = list(map(lambda city: [city, self.salaries_city_level[city]], self.salaries_city_level.keys()))
        rows3 = list(map(lambda city: [city, self.vacancies_city_count[city]], self.vacancies_city_count.keys()))

        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("pdf_template.html")
        pdf_template = template.render(graph_name='graph.png',
                                       vacancy_name=vacancy_name, headers1=headers1, headers2=headers2, headers3=headers3,
                                       rows1=rows1, rows2=rows2, rows3=rows3)
        config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
        options = {'enable-local-file-access': None}
        pdfkit.from_string(pdf_template, 'report.pdf', options=options, configuration=config)


#####################################################################################################################################


def get_vacancies():
    input_requests = ["Введите название файла: ", "Введите параметр фильтрации: ", "Введите параметр сортировки: ",
                      "Обратный порядок сортировки (Да / Нет): ", "Введите диапазон вывода: ",
                      "Введите требуемые столбцы: "]
    input_info = [input(input_request) for input_request in input_requests]
    normalize_result = normalize_input_info(input_info)
    if normalize_result != "Нормализация прошла успешно":
        return normalize_result
    (headers, info) = csv_reader(input_info[0])
    filtered_csv_data = csv_filter(headers, info)
    if len(filtered_csv_data) == 0:
        return "Нет данных"
    formatted_info = info_formatter(filtered_csv_data)
    filtered_info = info_filter(formatted_info, input_info[1])
    if len(filtered_info) == 0:
        return "Ничего не найдено"
    if input_info[2] != '№':
        filtered_info = info_sorter(filtered_info, input_info[2], input_info[3])
    print_vacancies(filtered_info, input_info[4], input_info[5])


def get_statistics():
    input_requests = ["Введите название файла: ", "Введите название профессии: "]
    input_info = [input(input_request) for input_request in input_requests]
    # input_info = ["vacancies_by_year.csv", "Аналитик"]
    if os.stat(input_info[0]).st_size == 0:
        print("Пустой файл")
        return
    data_set = DataSet(input_info[0])
    if len(data_set.vacancies_objects) == 0:
        print("Нет данных")
        return
    input_connect = InputConnect()
    formatted_info = input_connect.info_formatter(data_set.vacancies_objects)
    info = input_connect.info_finder(formatted_info, input_info[1])
    report = Report(info)
    report.generate_excel(input_info[1])
    report.generate_image(input_info[1])
    report.generate_pdf(input_info[1])


def main_function():
    main_input_request = "Выберите тип вывода: "
    main_input_info = input(main_input_request)
    # input_info = ["Вакансии"]
    if main_input_info != "Вакансии" and main_input_info != "Статистика":
        print("Введён неправильный тип вывода")
        return
    if main_input_info == "Вакансии":
        get_vacancies()
    else:
        get_statistics()


main_function()

# Hello, World

import operator
import multiprocessing
import time
import pdfkit
import requests
import csv
import json
import sqlite3
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import xml.etree.ElementTree as ET
from os import listdir, stat
from os.path import isfile, join
from functools import reduce, cmp_to_key
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from jinja2 import Environment, FileSystemLoader
from locale import atof, setlocale, LC_NUMERIC


class Vacancy:
    """Класс для представления вакансии

    Attributes:
        name (str): Название вакансии
        description (str):  Описание вакансии
        key_skills (list[str]): Необходимые скиллы для вакансии
        experience_id (str): Необходимый опыт для вакансии
        premium (str): Является ли вакансия премиумной
        employer_name (str): Название компании
        salary (str | float): Величина оклада
        area_name (str): Название города
        published_at (str): Время публикации вакансии
    """
    def __init__(self, name, description, key_skills, experience_id, premium,
                 employer_name, salary, area_name, published_at):
        """Инициализирует объект Vacancy

        Args:
            name (str):
            description (str | None):
            key_skills (list[str] | None):
            experience_id (str | None):
            premium (str | None):
            employer_name (str | None):
            salary (str | float):
            area_name (str):
            published_at (str):
        """
        self.name = name
        self.description = description
        self.key_skills = key_skills
        self.experience_id = experience_id
        self.premium = premium
        self.employer_name = employer_name
        self.salary = salary
        self.area_name = area_name
        self.published_at = published_at


# class Salary:
#     """Класс для представления оклада
#
#     Attributes:
#         salary_from (str):
#         salary_to (str):
#         salary_gross (str):
#         salary_currency (str):
#     """
#     def __init__(self, salary_from, salary_to, salary_gross, salary_currency):
#         """Инициализирует объект Salary
#
#         Args:
#             salary_from (str):
#             salary_to (str):
#             salary_gross (any):
#             salary_currency (str):
#         """
#         self.salary_from = salary_from
#         self.salary_to = salary_to
#         self.salary_gross = salary_gross
#         self.salary_currency = salary_currency
#
#     def currency_to_rur(self):
#         """Переводит верхнюю и нижнюю вилки оклада в рубли
#
#         Returns:
#             list[int,int]: Верхняя и нижняя вилки оклада в рублях
#         """
#         dic_currency_to_rur = {"AZN": 35.68, "BYR": 23.91, "EUR": 59.90, "GEL": 21.74,
#                                "KGS": 0.76, "KZT": 0.13, "RUR": 1, "UAH": 1.64, "USD": 60.66,
#                                "UZS": 0.0055}
#         # dic_currency_to_rub = {"Манаты": 35.68, "Белорусские рубли": 23.91, "Евро": 59.90, "Грузинский лари": 21.74,
#         #                        "Киргизский сом": 0.76, "Тенге": 0.13, "Рубли": 1, "Гривны": 1.64, "Доллары": 60.66,
#         #                        "Узбекский сум": 0.0055}
#         return [(int(x.replace(' ', '')) if x != '' else 0) * dic_currency_to_rur[self.salary_currency]
#                            for x in (self.salary_from, self.salary_to)]
#
#     def get_salary(self):
#         """Передаёт среднее значение оклада
#
#         Returns:
#             float: Среднее оклада
#         """
#         return sum(self.currency_to_rur()) / 2

class CurrencyApiConnect:
    """Класс для получения данных из внешних api и формировании по ним файлов

    Attributes:
        connect (sqlite3.connect): Объект управления базой данных
        cursor (sqlite3.connect): Объект управления базой данных
        db_path (str): Путь к базе данных
    """
    def __init__(self, db_path):
        """Инициализация объекта CurrencyApiConnect

        Args:
            db_path (str): Путь к базе данных
        """
        self.connect = sqlite3.connect(db_path)
        self.cursor = self.connect.cursor()
        self.db_path = db_path

    def get_currency_quotes(self, year_borders):
        """Получение обозначений валют и соответствующих им значений котировок по диапазону годов

        Args:
            year_borders (tuple[str, str]): Границы временного периода, с которого нужно получить котировки.

        Returns:
            dict[str: dict[str: float]]: Месяц и соответствующие котировки валют по месяцам
        """
        setlocale(LC_NUMERIC, 'French_Canada.1252')
        quotes_for_months = {}
        for year in range(int(year_borders[0]), int(year_borders[1]) + 1):
            for month in range(1, 13):
                month = format(month, '02d')
                req = requests.get(f"http://www.cbr.ru/scripts/XML_daily.asp?date_req=01/{month}/{year}")
                root_node = ET.ElementTree(ET.fromstring(req.text)).getroot()
                quotes = {tag.find('CharCode').text: atof(tag.find('Value').text) / atof(tag.find('Nominal').text)
                          for tag in root_node.findall('Valute')}
                quotes_for_months[f"{year}-{month}"] = quotes
                req.close()
                time.sleep(0.03)
        return quotes_for_months

    def save_currency_quotes_in_db(self, quotes_for_months, currencies):
        """Запись котировок валют в db файл

        Args:
            quotes_for_months (list[tuple[str, dict[str: float]]]): Котировки валют по месяцам
            currencies (list[str]): Названия валют
        """
        def db_query_join(query, type, join_str, data):
            i = 0
            while True:
                query = f"{query}{data[i]}{type}"
                if (i != len(data) - 1):
                    query += join_str
                    i += 1
                else:
                    query += ');'
                    break
            return query

        query = db_query_join("CREATE TABLE IF NOT EXISTS quotes(\ndate TEXT PRIMARY KEY,\n", ' REAL', ',\n', currencies)
        self.cursor.execute(query)
        self.connect.commit()
        query = db_query_join("INSERT INTO quotes\nVALUES(?, ", '', ', ', ['?' for _ in currencies])
        for (date, quotes_for_month) in quotes_for_months.items():
            quotes = tuple([date])
            quotes += tuple(list(map(lambda currency: quotes_for_month.get(currency, None)
                                                      if currency != 'RUR' else 1, currencies)))
            self.cursor.execute(query, quotes)
        self.connect.commit()


    def read_currency_quotes_from_db(self, currencies):
        """Чтение котировок валют из db файла

        Args:
            currencies (list[str]): Названия валют

        Returns:
            dict[str: dict[str: float]]: Котировки валют по годам
        """
        self.cursor.execute("SELECT * FROM quotes;")
        date_and_quotes = self.cursor.fetchall()
        quotes_for_years = {date_and_quotes[date_index][0]:
                                {currencies[currency_index]: date_and_quotes[date_index][currency_index + 1]
                                 for currency_index in range(len(currencies))}
                            for date_index in range(len(date_and_quotes))}
        return quotes_for_years


class HHruApiConnect:
    def save_vacancy_data_for_past_day(self):
        yesterday = time.strftime('%Y-%m-%d' , time.gmtime( time.time() - 86400 ))
        with open("vacancies_for_past_day.csv", mode="w", encoding='utf-8') as file:
            fileWriter = csv.writer(file, delimiter=",", lineterminator="\r")
            fileWriter.writerow(['name', 'salary_from', 'salary_to', 'salary_currency', 'area_name', 'published_at'])
            for time_from, time_to in (('00:00:00', '10:00:00'), ('10:00:00', '13:00:00'),
                                       ('13:00:00', '16:00:00'), ('16:00:00', '23:59:59')):
                for page in range(20):
                    vacancy_data = json.loads(self._get_vacancy_data_from_HHru(yesterday, time_from, time_to, page))
                    for item in vacancy_data['items']:
                        fileWriter.writerow([item['name'], item['salary']['from'], item['salary']['to'],
                                            item['salary']['currency'], item['area']['name'], item['published_at']])
                    time.sleep(0.05)

    def _get_vacancy_data_from_HHru(self, date, time_from, time_to, page):
        params = {
            'specialization': 1,
            'only_with_salary': True,
            'date_from': f'{date}T{time_from}',
            'date_to': f'{date}T{time_to}',
            'per_page': 100,
            'page': page
        }
        req = requests.get('https://api.hh.ru/vacancies', params)
        vacancy_data = req.content.decode()
        req.close()
        return vacancy_data


class DataSet:
    """Класс для получения информации из файла csv формата и базовой работы над данными из него

    """
    def split_csv_by_year(self, file_path):
        """Разделение csv файла по годам.

        Args:
            file_path (str): Путь к csv файлу
        """
        (headers, years_vacancy_info) = self._read_big_csv(file_path)
        popular_currencies = self._get_most_popular_currencies(years_vacancy_info)
        currency_db = CurrencyApiConnect('currency_quotes.db')
        quotes = currency_db.get_currency_quotes(self._get_year_borders(years_vacancy_info))
        currency_db.save_currency_quotes_in_db(quotes, popular_currencies)
        popular_currency_quotes = currency_db.read_currency_quotes_from_db(popular_currencies)

        filtered_years_vacancy_info = {}
        currency_to_rur = {"AZN": 35.68, "BYR": 23.91, "EUR": 59.90, "GEL": 21.74, "KGS": 0.76,
                           "KZT": 0.13, "RUR": 1, "UAH": 1.64, "USD": 60.66, "UZS": 0.0055}
        for year, year_info in years_vacancy_info.items():
            filtered_year_info = []
            for vacancy_info in year_info:
                if vacancy_info[3] not in popular_currencies \
                        or any(map(lambda x: x == '', (vacancy_info[0], vacancy_info[3], vacancy_info[-2], vacancy_info[-1]))): continue
                quote_value = popular_currency_quotes[vacancy_info[-1][:7]][vacancy_info[3]]
                salary = float(quote_value if quote_value != '' else currency_to_rur[vacancy_info[3]]) \
                         * (self._int_or_default(vacancy_info[1], 0) + self._int_or_default(vacancy_info[2], 0)) / 2
                if salary == 0: continue
                filtered_year_info.append([vacancy_info[0], salary, vacancy_info[4], vacancy_info[5]])
            filtered_years_vacancy_info[year] = filtered_year_info
        self._create_years_csv(['name', 'salary', 'area_name', 'published_at'], filtered_years_vacancy_info)

    def _read_big_csv(self, file_path):
        headers = []
        years_info = {}
        with open(file_path, encoding="utf-8-sig") as f:
            for row in map(lambda x: x, csv.reader(f)):
                if len(headers) == 0:
                    headers = row
                    continue
                if len(row) != len(headers):
                    continue
                year = row[-1][0:4]
                if year in years_info:
                    years_info[year].append(row)
                else:
                    years_info[year] = [row]
        return headers, years_info

    def _get_most_popular_currencies(self, years_vacancy_info):
        currency_count = {}
        for year_info in years_vacancy_info.values():
            for vacancy_info in year_info:
                if vacancy_info[3] not in currency_count:
                    currency_count[vacancy_info[3]] = 1
                else:
                    currency_count[vacancy_info[3]] += 1
        return [pair[0] for pair in currency_count.items() if pair[1] >= 5000]

    def _get_year_borders(self, years_vacancy_info):
        keys = list(years_vacancy_info.keys())
        return years_vacancy_info[keys[0]][0][-1][:4], years_vacancy_info[keys[-1]][0][-1][:4]

    def _int_or_default(self, value, default):
        dotIndex = value.find('.')
        dotIndex = dotIndex if dotIndex != -1 else len(value)
        return int(value[:dotIndex]) if value != '' else default

    def _create_years_csv(self, headers, years_vacancy_info):
        for year, info in years_vacancy_info.items():
            with open(f"years/{year}.csv", mode="w", encoding='utf-8-sig') as csv_year:
                file_writer = csv.writer(csv_year, delimiter=",", lineterminator="\r")
                file_writer.writerow(headers)
                file_writer.writerows(info)

    def get_vacancies_from_file(self, csv_year_file_path):
        """Чтение информации из csv файла определённого года и запись в список списков, в котором каждому внутреннему
            списку соответствует одна строка из файла

        Args:
            csv_year_file_path (str): Путь к csv файлу определённого года

        Returns:
            list[list[str]]: Форматированный список вакансий
        """
        info = self._read_csv(csv_year_file_path)[1:]
        return self._create_vacancies(info)

    def _read_csv(self, file_path):
        """Чтение информации из csv файла я запись в список списков, в котором каждому внутреннему списку
            соответствует одна строка из файла

        Args:
            file_path (str): Путь к csv файлу

        Returns:
            list[list[str]]: Форматированный список вакансий
        """
        with open(file_path, encoding="utf-8-sig") as f:
            reader_info = [x for x in csv.reader(f)]
            reader_info.pop(0)
            f.close()
        return reader_info

    def _create_vacancies(self, info):
        """Преобразование данных из csv файла в список вакансий, в котором каждой вакансии соответствует одна строка
            из файла (для статистики)

        Args:
            info (list[list[str]]): Строки csv файла

        Returns:
            list[Vacancy]: Форматированный список вакансий
        """
        return [Vacancy(info_row[0], None, None, None, None, None, info_row[1], info_row[2], info_row[3])
                for info_row in info]


class InputConnect:
    """Класс для работы над списком Vacancy: полное форматирование, нахождения необходимых вакансий

    """
    def info_formatter(self, vacancies):
        """Нормализация данных в вакансиях

        Args:
            vacancies (list[Vacancy]): Список вакансий

        Returns:
            list[Vacancy]: Результат форматирования
        """
        def formatter_published_at(attr_value):
            """Получение года из строки, содержащей дату

            Args:
                attr_value (str): Значение времени публикации вакансии

            Returns:
                str: Год публикации
            """
            return attr_value[0:4]

        for vacancy in vacancies:
            setattr(vacancy, "published_at", formatter_published_at(getattr(vacancy, "published_at")))
        return vacancies

    def year_info_finder(self, vacancies, finder_parameter):
        """Формирование информации по годам о вакансиях: уровень зарплат по годам, уровень зарплат по годам для
            выбранной вакансии, количество вакансий по годам, количество вакансий по годам для выбранной вакансии,
            уровень зарплат по городам, количество вакансий по городам, общее количество вакансий

        Args:
            vacancies (list[Vacancy]): Список вакансий
            finder_parameter (str): Название вакансии в качестве параметра фильтрации

        Returns:
            tuple[ dict[int: tuple[int, int]], dict[int: tuple[int, int]], dict[int: int], dict[int: int] ]: Группа
                списков
        """
        year = int(vacancies[0].published_at)
        salaries_year_level, selected_salary_year_level, vacancies_year_count, selected_vacancy_year_count, = \
            {}, {}, {}, {}
        for vacancy in vacancies:
            salary = float(vacancy.salary)
            if year not in salaries_year_level:
                salaries_year_level[year] = (salary, 1)
                vacancies_year_count[year] = 1
                selected_salary_year_level[year] = (0, 0)
                selected_vacancy_year_count[year] = 0
            else:
                sal_yr_lvl = salaries_year_level[year]
                salaries_year_level[year] = (sal_yr_lvl[0] + salary, sal_yr_lvl[1] + 1)
                vacancies_year_count[year] += 1
            if finder_parameter in vacancy.name:
                sel_sal_ye_lvl = selected_salary_year_level[year]
                selected_salary_year_level[year] = (sel_sal_ye_lvl[0] + salary, sel_sal_ye_lvl[1] + 1)
                selected_vacancy_year_count[year] += 1
        return self._year_info_calculating(salaries_year_level, selected_salary_year_level, vacancies_year_count,
                                           selected_vacancy_year_count)

    def city_info_finder(self, vacancies):
        """Формирование информации по годам о вакансиях: уровень зарплат по годам, уровень зарплат по годам для
            выбранной вакансии, количество вакансий по годам, количество вакансий по годам для выбранной вакансии,
            уровень зарплат по городам, количество вакансий по городам, общее количество вакансий

        Args:
            vacancies (list[Vacancy]): Список вакансий

        Returns:
            tuple[ dict[str: tuple[int, int]], dict[str: int] ]: Группа списков
        """
        salaries_city_level, vacancies_city_count = {}, {}
        for vacancy in vacancies:
            salary = float(vacancy.salary)
            if vacancy.area_name not in salaries_city_level:
                vacancies_city_count[vacancy.area_name] = 1
                salaries_city_level[vacancy.area_name] = (salary, 1)
            else:
                sal_ct_lvl = salaries_city_level[vacancy.area_name]
                salaries_city_level[vacancy.area_name] = (sal_ct_lvl[0] + salary, sal_ct_lvl[1] + 1)
                vacancies_city_count[vacancy.area_name] += 1
        return self._city_info_calculating(salaries_city_level, vacancies_city_count, len(vacancies))

    def _year_info_calculating(self, salaries_year_level, selected_salary_year_level, vacancies_year_count,
                               selected_vacancy_year_count):
        """Окончательное форматирование словарей, фильтрация, сортировка, выборка первого десятка для некоторых

        Args:
            salaries_year_level (dict[int: tuple[int, int]]): Уровень зарплат по годам
            selected_salary_year_level (dict[int: tuple[int, int]]): Уровень зарплат по годам для выбранной вакансии
            vacancies_year_count (dict[int: int]): Количество вакансий по годам
            selected_vacancy_year_count (dict[int: int]): Количество вакансий по годам для выбранной вакансии

        Returns:
            tuple[ dict[int: tuple[int, int]], dict[int: tuple[int, int]], dict[int: int], dict[int: int] ]:
                Уровень зарплат по годам, Уровень зарплат по годам для выбранной вакансии, Количество вакансий по годам,
                Количество вакансий по годам для выбранной вакансии
        """

        (salaries_year_level, selected_salary_year_level) = [
                {dict_pair[0]: int(dict_pair[1][0] / dict_pair[1][1]) if dict_pair[1][1] != 0 else int(dict_pair[1][0])
                    for dict_pair in dictionary.items()}
                for dictionary in (salaries_year_level, selected_salary_year_level)]

        return salaries_year_level, selected_salary_year_level, vacancies_year_count, selected_vacancy_year_count

    @staticmethod
    def _city_info_calculating(salaries_city_level, vacancies_city_count, vacancies_count):
        """Окончательное форматирование словарей, фильтрация, сортировка, выборка первого десятка для некоторых

        Args:
            salaries_city_level (dict[str: tuple[int, int]]): Уровень зарплат по городам
            vacancies_city_count (dict[str: int]): Количество вакансий по городам
            vacancies_count (int): Общее количество вакансий

        Returns:
            tuple[ dict[str: tuple[int, int]], dict[str: int] ]: Уровень зарплат по городам, Количество вакансий
                по городам
        """
        def sort_dict(dictionary):
            """Сортировка словаря лексикографически

            Args:
                dictionary (dict[str: int]): Словарь для сортировки

            Returns:
                tuple[dict[str: int], dict[str: int]]: Отсортированный словарь
            """
            dict_pairs = [(key, value) for key, value in dictionary.items()]
            dict_pairs.sort(key=cmp_to_key(lambda x, y: -1 if x[1] <= y[1] else 1))
            return dict(dict_pairs)

        vacancies_city_count = {dict_pair[0]: float(f"{dict_pair[1] / vacancies_count:.4f}")
                                for dict_pair in vacancies_city_count.items()}
        vacancies_city_count = {dict_pair[0]: dict_pair[1] for dict_pair in vacancies_city_count.items() if dict_pair[1] >= 0.01}
        vacancies_city_count = sort_dict(vacancies_city_count)
        vacancies_city_count = {dict_pair[0]: f"{round(dict_pair[1] * 100, 2)}%" for dict_pair in vacancies_city_count.items()}
        salaries_city_level = {dict_pair[0]:
                                int(dict_pair[1][0] / dict_pair[1][1]) if dict_pair[1][1] != 0 else int(dict_pair[1][0])
                               for dict_pair in salaries_city_level.items()}
        salaries_city_level = {dict_pair[0]: dict_pair[1] for dict_pair in salaries_city_level.items() if dict_pair[0] in vacancies_city_count}
        salaries_city_level = sort_dict(salaries_city_level)
        vacancies_city_count = {k: vacancies_city_count[k] for k in list(vacancies_city_count)[-10:][::-1]}
        salaries_city_level = {k: salaries_city_level[k] for k in list(salaries_city_level)[-10:][::-1]}
        return salaries_city_level, vacancies_city_count


class Report:
    """Класс для генерации файлов по анализу статистики: графиков, excel таблиц, общего pdf-файла

    Attributes:
        salaries_year_level (dict[int: tuple[int, int]]): Уровень зарплат по годам
        selected_salary_year_level (dict[int: tuple[int, int]]): Уровень зарплат по годам для выбранной вакансии
        vacancies_year_count (dict[int: int]): Количество вакансий по годам
        selected_vacancy_year_count (dict[int: int]): Количество вакансий по годам для выбранной вакансии
        salaries_city_level (dict[str: tuple[int, int]]): Уровень зарплат по городам
        vacancies_city_count (dict[str: int]): Количество вакансий по городам
    """
    def __init__(self, vacancy_info):
        """Инициализация объекта Report

        Args:
            vacancy_info (tuple[dict[int: tuple[int, int]], dict[int: tuple[int, int]], dict[int: int],
             dict[int: int], dict[str: tuple[int, int]], dict[str: int]]):
             Все словари созданные методом info_finder класса Input_Connect
        """
        self.salaries_year_level = vacancy_info[0]
        self.vacancies_year_count = vacancy_info[1]
        self.selected_salary_year_level = vacancy_info[2]
        self.selected_vacancy_year_count = vacancy_info[3]
        self.salaries_city_level = vacancy_info[4]
        self.vacancies_city_count = vacancy_info[5]

    def print_statistics(self):
        """Выводит на печать все статистику

        """
        print("Динамика уровня зарплат по годам:", self.salaries_year_level)
        print("Динамика количества вакансий по годам:", self.vacancies_year_count)
        print("Динамика уровня зарплат по годам для выбранной профессии:", self.selected_salary_year_level)
        print("Динамика количества вакансий по годам для выбранной профессии:", self.selected_vacancy_year_count)
        print("Уровень зарплат по городам (в порядке убывания):", self.salaries_city_level)
        print("Доля вакансий по городам (в порядке убывания):", self.vacancies_city_count)

    def generate_excel(self, vacancy_name):
        """Создание excel-файла основываясь на словарях аттрибутов объекта Report

        Args:
            vacancy_name (str): Название выбранной вакансии
        """
        wb = Workbook()
        stats_by_year = wb.worksheets[0]
        stats_by_year.title = "Cтатистика по годам"
        stats_by_city = wb.create_sheet("Cтатистика по городам")

        stats_by_year.append(["Год", "Средняя зарплата", f"Средняя зарплата - {vacancy_name}",
                              "Количество вакансий", f"Количество вакансий - {vacancy_name}"])
        for i, year in enumerate(self.salaries_year_level.keys(), 2):
            stats_by_year.cell(row=i, column=1, value=year)
            for j, dictionary in enumerate((self.salaries_year_level, self.vacancies_year_count,
                                            self.selected_salary_year_level, self.selected_vacancy_year_count), 2):
                stats_by_year.cell(row=i, column=j, value=dictionary[year])

        stats_by_city.append(["Город", "Уровень зарплат", "", "Город", "Доля вакансий"])
        for i, city in enumerate(self.salaries_city_level.keys(), 2):
            stats_by_city.cell(row=i, column=1, value=city)
            stats_by_city.cell(row=i, column=2, value=self.salaries_city_level[city])
        for i, city in enumerate(self.vacancies_city_count.keys(), 2):
            stats_by_city.cell(row=i, column=4, value=city)
            stats_by_city.cell(row=i, column=5, value=self.vacancies_city_count[city])

        self._slylize_wb(wb)
        wb.save('report.xlsx')

    @staticmethod
    def _slylize_wb(wb):
        """Стилизация рабочего листа excel-файла

        Args:
            wb (Workbook): Excel-лист
        """
        bold_font = Font(bold=True)
        thin = Side(border_style="thin", color="000000")
        outline = Border(top=thin, left=thin, right=thin, bottom=thin)
        for worksheet in wb.worksheets:
            for column_cells in worksheet.columns:
                length = max(len(str(cell.value) if cell.value is not None else "") for cell in column_cells)
                worksheet.column_dimensions[column_cells[0].column_letter].width = length + 3
            for cell in worksheet[1]:
                cell.font = bold_font
            for column in tuple(worksheet.columns):
                if column[1].value is None:
                    continue
                for cell in column:
                    cell.border = outline

    def generate_image(self, vacancy_name):
        """Создание графиков основываясь на словарях аттрибутов объекта Report

        Args:
            vacancy_name (str): Название выбранной вакансии
        """
        fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(2, 2, figsize=(12, 7.5), layout='constrained')
        self._generate_salary_year_levels_graph(ax1, vacancy_name)
        self._generate_vacancy_year_count_graph(ax2, vacancy_name)
        self._generate_salary_city_levels_graph(ax3)
        self._generate_vacancy_city_count_graph(ax4)
        plt.savefig('graph.png')

    def _generate_salary_year_levels_graph(self, ax, vacancy_name):
        """Создание графика уровня зарплат по годам

        Args:
            ax (Ax): Объект графика
            vacancy_name (str): Название выбранной вакансии
        """
        ax_labels = self.salaries_year_level.keys()
        x = np.arange(len(ax_labels))
        width = 0.35
        ax.bar(x - width / 2, self.salaries_year_level.values(), width, label='Средняя з/п')
        ax.bar(x + width / 2, self.selected_salary_year_level.values(), width, label=f'З/п {vacancy_name}')
        ax.set_xticks(x, ax_labels, fontsize=8, rotation=90, ha='right')
        ax.set_title("Уровень зарплат по годам")
        ax.yaxis.grid(True)
        ax.legend(fontsize=8, loc='upper left')

    def _generate_vacancy_year_count_graph(self, ax, vacancy_name):
        """Создание графика количества вакансий по годам

        Args:
            ax (Ax): Объект графика
            vacancy_name (str): Название выбранной вакансии
        """
        ax_labels = self.vacancies_year_count.keys()
        x = np.arange(len(ax_labels))
        width = 0.35
        ax.bar(x - width / 2, self.vacancies_year_count.values(), width, label='Количество вакансий')
        ax.bar(x + width / 2, self.selected_vacancy_year_count.values(), label=f'Количество вакансий {vacancy_name}')
        ax.set_xticks(x, ax_labels, fontsize=8, rotation=90, ha='right')
        ax.set_title("Количество вакансий по годам")
        ax.yaxis.grid(True)
        ax.legend(fontsize=8, loc='upper left')

    def _generate_salary_city_levels_graph(self, ax):
        """Создание графика уровня зарплат по городам

        Args:
            ax (Ax): Объект графика
        """
        ax_labels = self.salaries_city_level.keys()
        y_pos = np.arange(len(ax_labels))
        ax.barh(y_pos, self.salaries_city_level.values(), align='center')
        ax.set_yticks(y_pos, fontsize=8, labels=ax_labels)
        ax.invert_yaxis()
        ax.set_title("Уровень зарплат по городам")

    def _generate_vacancy_city_count_graph(self, ax):
        """Создание графика количества вакансий по городам

        Args:
            ax (Ax): Объект графика
        """
        ax_labels, values = list(self.vacancies_city_count.keys()), self.vacancies_city_count.values()
        ax_labels.append('Другие')
        values = list(map(lambda value: float(value[:-1]), values))
        values.append(100 - sum(values))
        ax.pie(values, labels=ax_labels)
        ax.set_title("Доля вакансий по городам")

    def generate_pdf(self, vacancy_name):
        """Создание pdf-файла основываясь на словарях аттрибутов объекта Report

        Args:
            vacancy_name (str): Название выбранной вакансии
        """
        headers1, headers2, headers3 = (["Год", "Средняя зарплата", f"Средняя зарплата - {vacancy_name}",
                                        "Количество вакансий", f"Количество вакансий - {vacancy_name}"],
                                        ["Город", "Уровень зарплат"], ["Город", "Доля вакансий"])
        rows1 = list(map(lambda year: [year] + [dictionary[year] for dictionary in
                                       (self.salaries_year_level, self.vacancies_year_count,
                                        self.selected_salary_year_level, self.selected_vacancy_year_count)]
                         , self.salaries_year_level.keys()))
        rows2 = list(map(lambda city: [city, self.salaries_city_level[city]], self.salaries_city_level.keys()))
        rows3 = list(map(lambda city: [city, self.vacancies_city_count[city]], self.vacancies_city_count.keys()))

        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("pdf_template.html")
        pdf_template = template.render(graph_name='graph.png',
                                       vacancy_name=vacancy_name, headers1=headers1, headers2=headers2,
                                       headers3=headers3, rows1=rows1, rows2=rows2, rows3=rows3)
        config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
        options = {'enable-local-file-access': None}
        pdfkit.from_string(pdf_template, 'report.pdf', options=options, configuration=config)


######################################################################################################################


class Consumer(multiprocessing.Process):
    """Служит для представления одного процесса, который берёт одну задачу из очереди задач и после выполнения кладёт
        результат в очереди результатов

    Attributes:
        task_queue (multiprocessing.JoinableQueue): Очередь задач
        results (multiprocessing.Queue): Очередь, куда будут складываться результаты
    """
    def __init__(self, task_queue, results):
        """Инициализация объекта Consumer

        Args:
            task_queue (multiprocessing.JoinableQueue): Очередь задач
            vacancies_info (multiprocessing.Queue): Очередь, куда будут складываться результаты
        """
        multiprocessing.Process.__init__(self)
        self.task_queue = task_queue
        self.results = results

    def run(self):
        """Выполняет одну задачу, полученную из списка задач, и сохраняет результат в соответствующих
            очередях результатов

        """
        while True:
            temp_task = self.task_queue.get()

            if temp_task is None:
                self.task_queue.task_done()
                break

            answer = temp_task.process()
            self.task_queue.task_done()
            self.results.put(answer)

class ReadTask():
    """Представляет собой одну задачу для выполнения процессом Consumer; читает данные из заданного csv файла и
        форматирует их в вакансии

    Attributes:
        file_name (str): Название файла, из которого нужно брать данные
        data_set (DadaSet): Объект DadaSet для анализа данных
        input_connect (InputConnect): Объект InputConnect для форматирования вакансий
    """
    def __init__(self, file_name, data_set, input_connect):
        """Инициализирует один объект класса Task

        Args:
            file_name (str): Название файла, из которого нужно брать данные
            input_connect (InputConnect): Объект InputConnect для форматирования вакансий
        """
        self.file_name = file_name
        self.data_set = data_set
        self.input_connect = input_connect

    def process(self):
        """Служит командой, которую нужно будет выполнять процессу Consumer

        Returns:
            list[Vacancy]: Список вакансий за соответствующий год и словари, содержащие статистику
        """
        vacancies = self.data_set.get_vacancies_from_file(self.file_name)
        formatted_vacancies = self.input_connect.info_formatter(vacancies)
        return formatted_vacancies


class CalculateTask():
    """Представляет собой одну задачу для выполнения процессом Consumer; составляет статистику о вакансиях
        по определённому году и заданной профессии

    Attributes:
        data_set (DadaSet): Объект DadaSet для анализа данных
        vacancies (list[Vacancy]): Вакансии для анализа
        input_connect (InputConnect): Объект InputConnect для форматирования и составления статистики по данным
    """
    def __init__(self, vacancy_name, vacancies, input_connect):
        """Инициализирует один объект класса Task

        Args:
            vacancy_name (str): Название вакансии для составления статистики
            vacancies (list[Vacancy]): Вакансии для анализа
            input_connect (InputConnect): Объект InputConnect для форматирования и составления статистики по данным
        """
        self.vacancy_name = vacancy_name
        self.vacancies = vacancies
        self.input_connect = input_connect

    def process(self):
        """Служит командой, которую нужно будет выполнять процессу Consumer

        Returns:
            tuple[list[Vacancy], tuple[dict[int: tuple[int, int]], dict[int: tuple[int, int]],
             dict[int: int], dict[int: int]]: Список вакансий за соответствующий год и словари, содержащие статистику
        """
        return self.input_connect.year_info_finder(self.vacancies, self.vacancy_name)


def get_statistics():
    """Получение информации с csv файла и создание графиков, таблиц и общего pdf-файл со статистикой
        на основе вводимых пользователем данных

    """
    def get_year_file_paths(folder_path):
        """Получение названия файлов из определённой папки

        Args:
            folder_path (str): Название папки, из которой нужно брать имена файлов

        Returns:
            list[str]: Названия файлов
        """
        return [f"{folder_path}/{file}" for file in listdir(folder_path) if isfile(join(folder_path, file))]

    def concat_dictionaries_in_tuples(tuples):
        """Используя группы словарей соединить каждый i-тый словарь

        Args:
            tuples (list[tuple[dict[int: int], dict[int: int], dict[int: int], dict[int: int]]]): Группы словарей для слияния

        Returns:
            tuple[dict[int: int], dict[int: int], dict[int: int], dict[int: int]]: Словари после слияния
        """
        (dict1, dict2, dict3, dict4) = {}, {}, {}, {}
        for dictionaries in tuples:
            dict1 = dict1 | dictionaries[0]
            dict2 = dict2 | dictionaries[1]
            dict3 = dict3 | dictionaries[2]
            dict4 = dict4 | dictionaries[3]
        return dict1, dict2, dict3, dict4

    def sort_dict_by_key(dictionary):
        """Сортировка словаря лексикографически по ключу

        Args:
            dictionary (dict[str: int]): Словарь для сортировки

        Returns:
            dict[str: int]: Отсортированный словарь
        """
        dict_pairs = [(key, value) for key, value in dictionary.items()]
        dict_pairs.sort(key=cmp_to_key(lambda x, y: -1 if x[0] <= y[0] else 1))
        return dict(dict_pairs)

    input_requests = ["Введите название файла: ", "Введите название профессии: "]
    # input_info = [input(input_request) for input_request in input_requests]
    input_info = ["vacancies_dif_currencies.csv", "Программист"]
    if stat(input_info[0]).st_size == 0:
        print("Пустой файл")
        return

    input_connect = InputConnect()
    data_set = DataSet()

    data_set.split_csv_by_year(input_info[0])
    year_file_paths = get_year_file_paths("years")

    tasks = multiprocessing.JoinableQueue()
    results = multiprocessing.Queue()
    consumers_count = multiprocessing.cpu_count() - 1

    consumers = [Consumer(tasks, results) for _ in range(consumers_count)]
    for consumer in consumers:
        consumer.start()
    for file_path in year_file_paths:
        tasks.put(ReadTask(file_path, data_set, input_connect))
    for _ in range(consumers_count):
        tasks.put(None)
    tasks.join()
    consumers.clear()

    all_vacancies_list = [results.get() for _ in range(len(year_file_paths))]
    tasks.empty()
    results.empty()

    tasks = multiprocessing.JoinableQueue()
    results = multiprocessing.Queue()
    consumers = [Consumer(tasks, results) for _ in range(consumers_count)]
    for consumer in consumers:
        consumer.start()
    for vacancies_list in all_vacancies_list:
        tasks.put(CalculateTask(input_info[1], vacancies_list, input_connect))
    for _ in range(consumers_count):
        tasks.put(None)
    tasks.join()
    consumers.clear()

    all_statistics = concat_dictionaries_in_tuples([results.get() for _ in range(len(year_file_paths))])
    year_statistics = tuple(sort_dict_by_key(dictionary) for dictionary in all_statistics)
    city_statistics = input_connect.city_info_finder(reduce(operator.concat, all_vacancies_list))
    report = Report(reduce(operator.concat, [year_statistics, city_statistics]))

    report.print_statistics()
    # report.generate_excel(input_info[1])
    # report.generate_image(input_info[1])
    # report.generate_pdf(input_info[1])